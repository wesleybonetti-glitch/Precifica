"""
Gerador de Excel - Módulo Nova Lei
Planilha completa de custos e formação de preços
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import datetime


class GeradorExcelNovaLei:
    """
    Gera planilha Excel completa com composição de custos
    """
    
    def __init__(self, empresa, cenario, resultado_calculo):
        self.empresa = empresa
        self.cenario = cenario
        self.resultado = resultado_calculo
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.ws.title = "Planilha de Custos"
        self.linha_atual = 1
        
        # Estilos
        self.estilo_titulo = Font(name='Arial', size=14, bold=True)
        self.estilo_subtitulo = Font(name='Arial', size=12, bold=True)
        self.estilo_cabecalho = Font(name='Arial', size=10, bold=True, color="FFFFFF")
        self.estilo_normal = Font(name='Arial', size=10)
        self.estilo_moeda = Font(name='Arial', size=10)
        
        self.preenchimento_cabecalho = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.preenchimento_subtotal = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        self.preenchimento_total = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        self.borda_fina = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def gerar(self):
        """Gera a planilha completa"""
        self.adicionar_cabecalho()
        self.adicionar_dados_proposta()
        self.adicionar_parametros()
        self.adicionar_detalhamento_postos()
        self.adicionar_resumo_financeiro()
        self.adicionar_bdi_citl()
        self.ajustar_largura_colunas()
        
        # Salvar em buffer
        buffer = BytesIO()
        self.wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    def adicionar_cabecalho(self):
        """Adiciona cabeçalho da planilha"""
        # Título
        self.ws.merge_cells(f'A{self.linha_atual}:F{self.linha_atual}')
        celula = self.ws[f'A{self.linha_atual}']
        celula.value = "PLANILHA DE CUSTOS E FORMAÇÃO DE PREÇOS"
        celula.font = self.estilo_titulo
        celula.alignment = Alignment(horizontal='center', vertical='center')
        self.linha_atual += 1
        
        # Subtítulo
        self.ws.merge_cells(f'A{self.linha_atual}:F{self.linha_atual}')
        celula = self.ws[f'A{self.linha_atual}']
        celula.value = "Modelo SEGES - IN 05/2017"
        celula.font = self.estilo_normal
        celula.alignment = Alignment(horizontal='center')
        self.linha_atual += 2
        
        # Dados da empresa
        if self.empresa:
            self.ws[f'A{self.linha_atual}'] = "Empresa:"
            self.ws[f'A{self.linha_atual}'].font = Font(bold=True)
            self.ws[f'B{self.linha_atual}'] = self.empresa.razao_social
            self.linha_atual += 1
            
            self.ws[f'A{self.linha_atual}'] = "CNPJ:"
            self.ws[f'A{self.linha_atual}'].font = Font(bold=True)
            self.ws[f'B{self.linha_atual}'] = self.empresa.cnpj
            self.linha_atual += 2
    
    def adicionar_dados_proposta(self):
        """Adiciona dados da proposta"""
        self.ws[f'A{self.linha_atual}'] = "DADOS DA PROPOSTA"
        self.ws[f'A{self.linha_atual}'].font = self.estilo_subtitulo
        self.linha_atual += 1
        
        dados = [
            ("Proposta ID:", self.cenario.id),
            ("Tipo de Serviço:", self.cenario.tipo_servico),
            ("Órgão:", self.cenario.razao_social or "Não informado"),
            ("Processo:", self.cenario.numero_processo or "Não informado"),
            ("Prazo de Execução:", f"{self.cenario.prazo_execucao or 0} meses"),
            ("Validade:", f"{self.cenario.validade_proposta or 0} dias"),
            ("Data:", datetime.now().strftime('%d/%m/%Y'))
        ]
        
        for label, valor in dados:
            self.ws[f'A{self.linha_atual}'] = label
            self.ws[f'A{self.linha_atual}'].font = Font(bold=True)
            self.ws[f'B{self.linha_atual}'] = valor
            self.linha_atual += 1
        
        self.linha_atual += 1
    
    def adicionar_parametros(self):
        """Adiciona parâmetros utilizados"""
        self.ws[f'A{self.linha_atual}'] = "PARÂMETROS UTILIZADOS"
        self.ws[f'A{self.linha_atual}'].font = self.estilo_subtitulo
        self.linha_atual += 1
        
        parametros = self.resultado.get('parametros_utilizados', {})
        
        dados_parametros = [
            ("INSS Patronal:", f"{parametros.get('inss_patronal', 0)}%"),
            ("Salário Educação:", f"{parametros.get('salario_educacao', 0)}%"),
            ("RAT/SAT:", f"{parametros.get('rat_sat', 0)}%"),
            ("FAP:", parametros.get('fap_multiplicador', 1.0)),
            ("Regime Tributário:", parametros.get('regime_tributario', 'Simples').upper()),
            ("Custos Indiretos:", f"{parametros.get('custos_indiretos_percentual', 0)}%"),
            ("Lucro:", f"{parametros.get('lucro_percentual', 0)}%")
        ]
        
        for label, valor in dados_parametros:
            self.ws[f'A{self.linha_atual}'] = label
            self.ws[f'A{self.linha_atual}'].font = Font(bold=True)
            self.ws[f'B{self.linha_atual}'] = valor
            self.linha_atual += 1
        
        self.linha_atual += 1
    
    def adicionar_detalhamento_postos(self):
        """Adiciona detalhamento por posto"""
        self.ws[f'A{self.linha_atual}'] = "DETALHAMENTO POR POSTO"
        self.ws[f'A{self.linha_atual}'].font = self.estilo_subtitulo
        self.linha_atual += 1
        
        # Cabeçalho da tabela
        colunas = ['Cargo', 'Qtd', 'Remuneração', 'Encargos', 'Provisões', 'Reposição', 'Insumos', 'Subtotal', 'CITL', 'Total']
        
        for col_idx, coluna in enumerate(colunas, start=1):
            celula = self.ws.cell(row=self.linha_atual, column=col_idx)
            celula.value = coluna
            celula.font = self.estilo_cabecalho
            celula.fill = self.preenchimento_cabecalho
            celula.alignment = Alignment(horizontal='center', vertical='center')
            celula.border = self.borda_fina
        
        self.linha_atual += 1
        
        # Dados dos postos
        postos = self.resultado.get('postos', [])
        
        for posto in postos:
            valores = [
                posto['cargo'],
                posto['quantidade_postos'],
                posto['modulo_1_remuneracao']['total'],
                posto['modulo_2_encargos_beneficios']['total'],
                posto['modulo_3_provisoes']['total'],
                posto['modulo_4_reposicao']['total'],
                posto['modulo_5_insumos']['total'],
                posto['subtotal_antes_citl'],
                posto['modulo_6_citl']['total'],
                posto['total_posto']
            ]
            
            for col_idx, valor in enumerate(valores, start=1):
                celula = self.ws.cell(row=self.linha_atual, column=col_idx)
                
                if isinstance(valor, (int, float)) and col_idx > 2:
                    celula.value = valor
                    celula.number_format = 'R$ #,##0.00'
                else:
                    celula.value = valor
                
                celula.border = self.borda_fina
                celula.alignment = Alignment(horizontal='right' if col_idx > 2 else 'left')
            
            self.linha_atual += 1
        
        self.linha_atual += 1
    
    def adicionar_resumo_financeiro(self):
        """Adiciona resumo financeiro"""
        self.ws[f'A{self.linha_atual}'] = "RESUMO FINANCEIRO"
        self.ws[f'A{self.linha_atual}'].font = self.estilo_subtitulo
        self.linha_atual += 1
        
        total = self.resultado.get('total_contrato', {})
        
        dados_resumo = [
            ("Remuneração Total", total.get('total_remuneracao', 0)),
            ("Encargos e Benefícios", total.get('total_encargos_beneficios', 0)),
            ("Provisões", total.get('total_provisoes', 0)),
            ("Reposição", total.get('total_reposicao', 0)),
            ("Insumos", total.get('total_insumos', 0)),
        ]
        
        for descricao, valor in dados_resumo:
            self.ws[f'A{self.linha_atual}'] = descricao
            self.ws[f'A{self.linha_atual}'].font = Font(bold=True)
            self.ws[f'B{self.linha_atual}'] = valor
            self.ws[f'B{self.linha_atual}'].number_format = 'R$ #,##0.00'
            self.linha_atual += 1
        
        # Subtotal
        self.ws[f'A{self.linha_atual}'] = "SUBTOTAL (antes de BDI/CITL)"
        self.ws[f'A{self.linha_atual}'].font = Font(bold=True)
        self.ws[f'A{self.linha_atual}'].fill = self.preenchimento_subtotal
        self.ws[f'B{self.linha_atual}'] = total.get('subtotal_antes_citl', 0)
        self.ws[f'B{self.linha_atual}'].number_format = 'R$ #,##0.00'
        self.ws[f'B{self.linha_atual}'].fill = self.preenchimento_subtotal
        self.linha_atual += 2
    
    def adicionar_bdi_citl(self):
        """Adiciona BDI/CITL"""
        self.ws[f'A{self.linha_atual}'] = "BDI / CITL"
        self.ws[f'A{self.linha_atual}'].font = self.estilo_subtitulo
        self.linha_atual += 1
        
        total = self.resultado.get('total_contrato', {})
        
        dados_bdi = [
            ("Custos Indiretos", total.get('custos_indiretos', 0)),
            ("Tributos", total.get('tributos', 0)),
            ("Lucro", total.get('lucro', 0)),
        ]
        
        for descricao, valor in dados_bdi:
            self.ws[f'A{self.linha_atual}'] = descricao
            self.ws[f'A{self.linha_atual}'].font = Font(bold=True)
            self.ws[f'B{self.linha_atual}'] = valor
            self.ws[f'B{self.linha_atual}'].number_format = 'R$ #,##0.00'
            self.linha_atual += 1
        
        self.linha_atual += 1
        
        # Total geral
        self.ws[f'A{self.linha_atual}'] = "VALOR TOTAL DA PROPOSTA"
        self.ws[f'A{self.linha_atual}'].font = Font(bold=True, size=12, color="FFFFFF")
        self.ws[f'A{self.linha_atual}'].fill = self.preenchimento_total
        self.ws[f'B{self.linha_atual}'] = total.get('total_contrato', 0)
        self.ws[f'B{self.linha_atual}'].number_format = 'R$ #,##0.00'
        self.ws[f'B{self.linha_atual}'].font = Font(bold=True, size=12, color="FFFFFF")
        self.ws[f'B{self.linha_atual}'].fill = self.preenchimento_total
    
    def ajustar_largura_colunas(self):
        """Ajusta largura das colunas"""
        larguras = {
            'A': 30,
            'B': 15,
            'C': 15,
            'D': 15,
            'E': 15,
            'F': 15,
            'G': 15,
            'H': 15,
            'I': 15,
            'J': 15
        }
        
        for coluna, largura in larguras.items():
            self.ws.column_dimensions[coluna].width = largura
